#!/usr/bin/env python3
"""Calibrate the Live Results confidence sigmas against VIC 2022 booth data.

The live election-night model (webapp/src/live/confidence.js) converts a
partially counted seat's projected 2CP margin into a win probability using a
count-driven sigma:

    sigma(f) = SIGMA_FLOOR + SIGMA_SCALE*sqrt(1-f) + LATE_SWING_SIGMA*(1-f)

where f is the counted fraction. The original constants were first estimates.
This script replaces guesswork with an empirical fit: it replays the 2022
Victorian election from booth-level 2CP data (VIC-2022-LA-2CP-Pollingplace.xlsx,
in Git LFS under data/raw/vic/202211/) and measures how far the running 2CP
share sits from the final result at each counted fraction.

Simulation (per district, many trials):
  1. Ordinary election-day booths are counted one by one in random order —
     mirroring the evening, when booth completion order is effectively random.
  2. The late categories are then added in the order they realistically land:
     Early Vote (prepoll centres, late evening), then Postal / Absent /
     Provisional / Marked as Voted over the following days.
  3. At each step, error = running 2CP share of the reference candidate minus
     the final share (pp). Errors are bucketed by counted fraction; the
     bucket standard deviation is the empirical sigma(f).

The sigma curve is then least-squares fitted (non-negative coefficients) to
FLOOR + SCALE*sqrt(1-f) + LATE*(1-f), separately for classic ALP-vs-Coalition
seats and for everything else (Greens/Independent finals), matching the two
families in confidence.js. The statewide correlated term is anchored to the
vote-weighted systematic skew between ordinary booths and the full count
(the "postals break differently" risk that early projections can't see).

Run:  python scripts/calibrate_live_sigma.py
      (requires the LFS Excel to be pulled; prints fitted constants)

The fitted values are hand-copied into webapp/src/live/confidence.js — keep
the constants there in sync with this script's output and note the seed.
"""
from __future__ import annotations

import random
from collections import defaultdict
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX = REPO_ROOT / "data" / "raw" / "vic" / "202211" / "VIC-2022-LA-2CP-Pollingplace.xlsx"

SEED = 20221126
TRIALS = 60                     # random booth orderings per district
BUCKETS = 20                    # counted-fraction buckets (5 % wide)

COALITION = {"LIB", "NAT", "LNP"}
# Late-count categories in realistic landing order. Early Vote arrives on the
# night (after most ordinary booths); the rest trickle in over following days.
LATE_ORDER = ["Early Vote", "Postal Vote", "Absent", "Provisional", "Marked as Voted"]
PSEUDO_CANDIDATES = {"MIS"}     # Mis-sorts rows are not candidate votes


def load_booth_2cp():
    """-> {district: {"booths": {pp_name: {party: votes}}, "parties": [p1, p2]}}"""
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(c) for c in next(rows)]
    ix = {name: header.index(name) for name in
          ("district_name", "pp_name", "party_code", "votes")}

    districts: dict[str, dict] = {}
    for row in rows:
        district = row[ix["district_name"]]
        party = row[ix["party_code"]]
        if district is None or party in PSEUDO_CANDIDATES:
            continue
        booth = str(row[ix["pp_name"]])
        votes = int(row[ix["votes"]] or 0)
        d = districts.setdefault(district, {"booths": {}, "parties": []})
        if party not in d["parties"]:
            d["parties"].append(party)
        d["booths"].setdefault(booth, defaultdict(int))[party] += votes
    return districts


def is_alp_coal(parties: list[str]) -> bool:
    return "ALP" in parties and any(p in COALITION for p in parties)


def simulate_district(d: dict, rng: random.Random):
    """Yield (counted_fraction, error_pp) points across TRIALS orderings."""
    parties = d["parties"]
    if len(parties) != 2:
        return
    ref = parties[0]
    booths = d["booths"]
    ordinary = [b for b in booths if b not in LATE_ORDER]
    late = [b for b in LATE_ORDER if b in booths]

    tot = {p: sum(bv.get(p, 0) for bv in booths.values()) for p in parties}
    total_votes = sum(tot.values())
    if total_votes == 0:
        return
    final_share = 100.0 * tot[ref] / total_votes

    for _ in range(TRIALS):
        order = ordinary[:]
        rng.shuffle(order)
        order += late                     # late categories land last, in order
        run_ref = run_tot = 0
        for booth in order:
            bv = booths[booth]
            run_ref += bv.get(ref, 0)
            run_tot += sum(bv.get(p, 0) for p in parties)
            if run_tot == 0:
                continue
            f = run_tot / total_votes
            err = 100.0 * run_ref / run_tot - final_share
            yield f, err


def bucket_sigmas(points):
    """[(f, err)] -> [(f_mid, sigma, n)] over BUCKETS equal-width buckets."""
    sums = defaultdict(float)
    sqs = defaultdict(float)
    ns = defaultdict(int)
    for f, err in points:
        b = min(BUCKETS - 1, int(f * BUCKETS))
        sums[b] += err
        sqs[b] += err * err
        ns[b] += 1
    out = []
    for b in sorted(ns):
        n = ns[b]
        if n < 30:
            continue
        mean = sums[b] / n
        var = max(0.0, sqs[b] / n - mean * mean)
        out.append(((b + 0.5) / BUCKETS, var ** 0.5, n))
    return out


MIN_FLOOR = 0.25   # recount / check-count / data-entry risk that a booth-order
                   # simulation cannot see — never let sigma collapse below this.
F_MIN = 0.02       # clamp for the sampling term at f -> 0 (caps the spike)


def fit_sigma_curve(buckets):
    """Weighted non-negative LS of sigma ~ a + b*sqrt((1-f)/max(f,F_MIN)) + c*(1-f).

    The middle basis is the finite-population sampling shape — the empirical
    booth-replay curve tracks sqrt((1-f)/f) almost exactly (the error of a
    partial count scales with sqrt(remaining/counted)), which the previous
    sqrt(1-f) basis cannot reproduce at small f.

    a is constrained to >= MIN_FLOOR. Weights are sqrt(n): points within one
    trial are autocorrelated, so raw n badly over-weights the near-100% buckets
    every trial passes through. Non-negativity via subset enumeration (3 coefs
    -> 8 subsets), each solved exactly with numpy — no scipy needed.
    """
    import numpy as np

    f = np.array([b[0] for b in buckets])
    y = np.array([b[1] for b in buckets]) - MIN_FLOOR
    w = np.sqrt(np.array([b[2] for b in buckets], dtype=float))
    X = np.column_stack([np.ones_like(f),
                         np.sqrt((1 - f) / np.maximum(f, F_MIN)),
                         (1 - f)])

    best = None
    for mask in range(8):
        keep = [j for j in range(3) if mask & (1 << j)]
        coef = np.zeros(3)
        if keep:
            Xk = X[:, keep] * w[:, None]
            ck, *_ = np.linalg.lstsq(Xk, y * w, rcond=None)
            if (ck < 0).any():
                continue
            coef[keep] = ck
        sse = float((((X @ coef - y) * w) ** 2).sum())
        if best is None or sse < best[0]:
            best = (sse, coef)
    a, b, c = best[1]
    return a + MIN_FLOOR, b, c


def main():
    rng = random.Random(SEED)
    districts = load_booth_2cp()
    print(f"loaded {len(districts)} districts from {XLSX.name}")

    pts_alp_coal, pts_other = [], []
    skews = []                            # (systematic ordinary-vs-final skew, weight)
    for name, d in sorted(districts.items()):
        if len(d["parties"]) != 2:
            print(f"  ! {name}: {len(d['parties'])} 2CP parties — skipped")
            continue
        pts = list(simulate_district(d, rng))
        (pts_alp_coal if is_alp_coal(d["parties"]) else pts_other).extend(pts)

        # Systematic skew: ALP-perspective share among ordinary booths only vs
        # the full count. This is what a "100 % of election-day booths, no
        # late votes yet" projection would be off by — the correlated risk.
        if is_alp_coal(d["parties"]):
            ref = "ALP"
            booths = d["booths"]
            parties = d["parties"]
            ord_ref = ord_tot = fin_ref = fin_tot = 0
            for booth, bv in booths.items():
                v_ref = bv.get(ref, 0)
                v_tot = sum(bv.get(p, 0) for p in parties)
                fin_ref += v_ref
                fin_tot += v_tot
                if booth not in LATE_ORDER:
                    ord_ref += v_ref
                    ord_tot += v_tot
            if ord_tot and fin_tot:
                skews.append((100 * ord_ref / ord_tot - 100 * fin_ref / fin_tot,
                              fin_tot))

    print(f"simulated points: ALP-vs-Coalition {len(pts_alp_coal):,}, "
          f"other finals {len(pts_other):,}")

    print("\nALP-vs-Coalition sigma by counted fraction:")
    b_ac = bucket_sigmas(pts_alp_coal)
    for f, s, n in b_ac:
        print(f"  f={f:4.2f}  sigma={s:5.2f}pp  (n={n:,})")
    floor, scale, late = fit_sigma_curve(b_ac)
    print(f"fit: SIGMA_FLOOR={floor:.2f}  SAMPLE_SIGMA={scale:.2f}  "
          f"LATE_SWING_SIGMA={late:.2f}   "
          f"[sigma = FLOOR + SAMPLE*sqrt((1-f)/max(f,{F_MIN})) + LATE*(1-f)]")
    for f, s, n in b_ac:
        pred = floor + scale * ((1 - f) / max(f, F_MIN)) ** 0.5 + late * (1 - f)
        print(f"    f={f:4.2f}  empirical={s:5.2f}  fitted={pred:5.2f}")

    print("\nNon-ALP/Coalition sigma by counted fraction:")
    b_ot = bucket_sigmas(pts_other)
    for f, s, n in b_ot:
        print(f"  f={f:4.2f}  sigma={s:5.2f}pp  (n={n:,})")
    o_floor, o_scale, o_late = fit_sigma_curve(b_ot)
    print(f"fit: NON_ALP_COAL_FLOOR={o_floor:.2f}  NON_ALP_COAL_SAMPLE={o_scale:.2f}  "
          f"NON_ALP_COAL_LATE={o_late:.2f}")
    for f, s, n in b_ot:
        pred = o_floor + o_scale * ((1 - f) / max(f, F_MIN)) ** 0.5 + o_late * (1 - f)
        print(f"    f={f:4.2f}  empirical={s:5.2f}  fitted={pred:5.2f}")

    wtot = sum(w for _, w in skews)
    mean_skew = sum(s * w for s, w in skews) / wtot
    var_skew = sum((s - mean_skew) ** 2 * w for s, w in skews) / wtot
    print(f"\nStatewide systematic ordinary-booths-vs-final ALP skew: "
          f"{mean_skew:+.2f}pp (cross-district spread {var_skew ** 0.5:.2f}pp)")
    print("CORR_BASE should cover this correlated skew at low counted "
          f"fractions: |skew| + spread ≈ {abs(mean_skew) + var_skew ** 0.5:.2f}pp")


if __name__ == "__main__":
    main()
